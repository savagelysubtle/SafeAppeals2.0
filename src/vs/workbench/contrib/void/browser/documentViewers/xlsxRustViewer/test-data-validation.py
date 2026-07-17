"""
Creates a test XLSX file with every type of data validation
so you can open it in the XLSX Rust Viewer and verify all features.

Run:  python test-data-validation.py
Output: test-data-validation.xlsx  (in the same folder)

Requires openpyxl:  pip install openpyxl
"""
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

wb = Workbook()
ws = wb.active
ws.title = "Validation Tests"

# --- Header styling ---
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")

def header(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = header_fill
    cell.font = header_font
    return cell

def label(ws, row, col, text):
    ws.cell(row=row, column=col, value=text)

# ── Column widths ──────────────────────────────────────────────
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 40

# ─────────────────────────────────────────────────────────────
# 1. Dropdown list (inline)
# ─────────────────────────────────────────────────────────────
header(ws, 1, 1, "Feature")
header(ws, 1, 2, "Try typing here →")
header(ws, 1, 3, "Description")

label(ws, 2, 1, "1. Dropdown list")
label(ws, 2, 3, 'Only "Yes", "No", or "Maybe" allowed. Click the arrow.')
dv_list = DataValidation(
    type="list",
    formula1='"Yes,No,Maybe"',
    showDropDown=False,   # False = SHOW the dropdown arrow
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Pick a value",
    prompt="Select Yes, No, or Maybe from the dropdown.",
    errorTitle="Invalid input",
    error='Please select "Yes", "No", or "Maybe".',
    errorStyle="stop",
)
dv_list.sqref = "B2"
ws.add_data_validation(dv_list)

# ─────────────────────────────────────────────────────────────
# 2. Whole number (between 1 and 10)
# ─────────────────────────────────────────────────────────────
label(ws, 3, 1, "2. Whole number (1–10)")
label(ws, 3, 3, "Only integers from 1 to 10. Try typing 0 or 11.")
dv_int = DataValidation(
    type="whole",
    operator="between",
    formula1="1",
    formula2="10",
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Rating",
    prompt="Enter a whole number between 1 and 10.",
    errorTitle="Out of range",
    error="The value must be an integer between 1 and 10.",
    errorStyle="stop",
)
dv_int.sqref = "B3"
ws.add_data_validation(dv_int)

# ─────────────────────────────────────────────────────────────
# 3. Decimal (0.0 to 5.0)
# ─────────────────────────────────────────────────────────────
label(ws, 4, 1, "3. Decimal (0.0–5.0)")
label(ws, 4, 3, "Decimals from 0.0 to 5.0. Try typing 5.1.")
dv_dec = DataValidation(
    type="decimal",
    operator="between",
    formula1="0",
    formula2="5",
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Score",
    prompt="Enter a decimal value from 0.0 to 5.0.",
    errorTitle="Invalid score",
    error="Score must be a decimal number between 0.0 and 5.0.",
    errorStyle="warning",   # Warning = shows Yes/No (doesn't hard-block)
)
dv_dec.sqref = "B4"
ws.add_data_validation(dv_dec)

# ─────────────────────────────────────────────────────────────
# 4. Date (>= 2020-01-01)
# ─────────────────────────────────────────────────────────────
label(ws, 5, 1, "4. Date (≥ 2020-01-01)")
label(ws, 5, 3, "Only dates on or after 2020-01-01.")
dv_date = DataValidation(
    type="date",
    operator="greaterThanOrEqual",
    formula1="2020-01-01",
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Date",
    prompt="Enter a date on or after 1 Jan 2020.",
    errorTitle="Too early",
    error="Date must be 2020-01-01 or later.",
    errorStyle="stop",
)
dv_date.sqref = "B5"
ws.add_data_validation(dv_date)

# ─────────────────────────────────────────────────────────────
# 5. Text length (5–20 characters)
# ─────────────────────────────────────────────────────────────
label(ws, 6, 1, "5. Text length (5–20 chars)")
label(ws, 6, 3, "Text must be 5–20 characters long.")
dv_len = DataValidation(
    type="textLength",
    operator="between",
    formula1="5",
    formula2="20",
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Description",
    prompt="Enter text between 5 and 20 characters.",
    errorTitle="Wrong length",
    error="Text must be between 5 and 20 characters.",
    errorStyle="information",  # Information = just informs, always accepts
)
dv_len.sqref = "B6"
ws.add_data_validation(dv_len)

# ─────────────────────────────────────────────────────────────
# 6. Custom formula (must be UPPERCASE)
# ─────────────────────────────────────────────────────────────
label(ws, 7, 1, "6. Custom formula (uppercase only)")
label(ws, 7, 3, "Cell must equal its own uppercase version. Try typing 'hello'.")
dv_custom = DataValidation(
    type="custom",
    formula1="=EXACT(B7,UPPER(B7))",
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Uppercase only",
    prompt="Type text in UPPERCASE only.",
    errorTitle="Not uppercase",
    error="The value must be in uppercase letters.",
    errorStyle="stop",
)
dv_custom.sqref = "B7"
ws.add_data_validation(dv_custom)

# ─────────────────────────────────────────────────────────────
# 7. Input message only (no restriction)
# ─────────────────────────────────────────────────────────────
label(ws, 8, 1, "7. Input message only (any value)")
label(ws, 8, 3, "No restriction, but shows a tooltip when you click the cell.")
dv_msg = DataValidation(
    type="any",
    showInputMessage=True,
    showErrorMessage=False,
    promptTitle="Hint",
    prompt="You can type anything here, but we recommend entering your full name.",
)
dv_msg.sqref = "B8"
ws.add_data_validation(dv_msg)

# ─────────────────────────────────────────────────────────────
# 8. Multi-cell dropdown applied to a range
# ─────────────────────────────────────────────────────────────
label(ws, 10, 1, "8. Multi-cell dropdown (B11:B15)")
label(ws, 10, 3, "Five cells, each with the same list: Low/Medium/High/Critical")
for r in range(11, 16):
    label(ws, r, 3, "← pick a priority")
dv_multi = DataValidation(
    type="list",
    formula1='"Low,Medium,High,Critical"',
    showDropDown=False,
    showInputMessage=True,
    showErrorMessage=True,
    promptTitle="Priority",
    prompt="Choose a priority level.",
    errorTitle="Invalid",
    error="Please choose Low, Medium, High, or Critical.",
    errorStyle="stop",
)
dv_multi.sqref = "B11:B15"
ws.add_data_validation(dv_multi)

# ─────────────────────────────────────────────────────────────
# Pre-fill one invalid value so "Circle Invalid Data" can be tested
# ─────────────────────────────────────────────────────────────
ws["B3"] = 999   # Whole number rule is 1–10; this is invalid → should get circled
ws["B4"] = 99.9  # Decimal rule is 0–5; this is invalid → should get circled

# ─────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────
out_path = "test-data-validation.xlsx"
wb.save(out_path)
print(f"✅  Saved:  {out_path}")
print()
print("Open it in the XLSX Rust Viewer and test:")
print("  Row 2  → click cell B2 → dropdown arrow should appear")
print("  Row 3  → B3 has value 999 (invalid) → try Circle Invalid Data button")
print("  Row 4  → B4 has value 99.9 (invalid) → try Circle Invalid Data button")
print("  Row 8  → click B8 → input message tooltip should appear")
print("  Rows 11-15 → each cell has a dropdown")
